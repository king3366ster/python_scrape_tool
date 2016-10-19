#! /usr/bin/env python
#-*- coding:utf-8 -*-
import os, pdb
import MySQLdb
from DBUtils.PooledDB import PooledDB
import openpyxl
from openpyxl.utils import get_column_letter

class SaveData:
    def __init__(self, path = '', logPath = ''):
        if path == '':
            pass
        elif not os.path.exists(path):
            os.makedirs(path)
        self.path = path

    def excelWriter(self, df, tb_name = '1', if_exists = 'replace', sheet_name = 'Sheet1', encoding = 'utf-8', id_offset = 0):
        if if_exists == 'append' or if_exists == 'append_ignore':
            try:
                wb = openpyxl.load_workbook('%s%s.xlsx' % (self.path, tb_name))
                ws = wb.get_sheet_by_name(sheet_name)
            except Exception as what:
                wb = openpyxl.workbook.Workbook()
                ws = wb.create_sheet(sheet_name, 0)
        elif if_exists == 'replace':
            wb = openpyxl.workbook.Workbook()
            ws = wb.create_sheet(sheet_name, 0)
        else:
            wb = openpyxl.workbook.Workbook()
            ws = wb.create_sheet(sheet_name, 0)

        dfIndexs = df['indexs']
        dfColumns = df['columns']
        dfValues = df['values']
        # reset columns
        for i in range(0, len(dfColumns)):
            index_col = i + 1
            index_row = 1
            col = get_column_letter(index_col + 1)
            ws.cell( '%s%s' %(col, index_row)).value = dfColumns[i]

        for i in range(0, len(dfIndexs)):
            index_row = dfIndexs[i] + 1 - id_offset
            ws.cell('A'+str(index_row)).value = dfIndexs[i] + id_offset
            for j in range(0, len(dfColumns)):
                index_col = j + 1
                col = get_column_letter(index_col + 1)
                tmpValue = dfValues[i][j]
                try:
                    if (ws.cell('%s%s' %(col, index_row)).value is not None) and ws.cell('%s%s' %(col, index_row)).value != '':
                        if if_exists == 'append_ignore':
                            pass
                        elif if_exists == 'append' or if_exists == 'replace':
                            ws.cell( '%s%s' %(col, index_row)).value = str(tmpValue)
                    else:
                        ws.cell( '%s%s' %(col, index_row)).value = tmpValue
                    # ws['%s%s' %(col, index_row)] = int(tmpValue)
                except Exception as what:
                    pass
                    # # print 'encoding error: ', col, index_row, what
                    # self.logger.error('write excel raw error: ' + str(col) + ' ' + str(index_row) + ' |_| reading error ' + str(what))
        wb.save(filename = '%s%s.xlsx' % (self.path, unicode(tb_name)))

    def mysqlConnect(self, sqlCfg):
        if 'host' in sqlCfg:
            _host = sqlCfg['host']
        else:
            _host = '127.0.0.1'
        self._host = _host

        if 'user' in sqlCfg:
            _user = sqlCfg['user']
        else:
            _user = 'root'
        self._user = _user

        if 'pwd' in sqlCfg:
            _pwd = sqlCfg['pwd']
        else:
            _pwd = ''
        self._pwd = _pwd

        if 'db' in sqlCfg:
            _db = sqlCfg['db']
        self._db = _db

        if 'port' in sqlCfg:
            _port = sqlCfg['port']
        else:
            _port = 3306
        self._port = _port

        try:
            # self.conn = MySQLdb.connect(host = _host, user = _user, passwd = _pwd, port = _port, charset = 'utf8', use_unicode = True)
            self.pool = PooledDB(MySQLdb, 5, host = _host, user = _user, passwd = _pwd, port = _port, db = _db, charset = 'utf8', use_unicode = True)
            # self.conn = pool.connection()
            # self.cursor = self.conn.cursor()
            # self.conn.select_db(_db)
            print 'Mysql connected %s %s' % (_host, _db)
        except MySQLdb.Error,e:
            print 'Mysql Error %d: %s' % (e.args[0], e.args[1])

    def createConn(self):
        self.conn = self.pool.connection()
        self.cursor = self.conn.cursor()

    def releaseConn(self):
        self.cursor.close()
        self.conn.close()

    def mysqlWriter(self, df, tb_name = 'test', if_exists = 'append', need_datetime = True, unique_key = None):
        if if_exists == 'replace':
            self.cursor.execute('DELETE FROM %s;' % tb_name)
            self.conn.commit()

        columns = df['columns']
        values = df['values']
        columns = map(lambda x: MySQLdb.escape_string(x), columns)

        if if_exists == 'append' or if_exists == 'replace':
            self.createConn()
            values_occup = ['%s'] * len(columns)

            # upsert_string = 'INSERT INTO %s (%s) VALUES (%s) ON DUPLICATE KEY UPDATE' % (tb_name, ','.join(columns), ','.join(values_occup))

            insert_string = 'INSERT INTO %s (%s) VALUES (%s);' % (tb_name, ','.join(columns), ','.join(values_occup))

            if unique_key is not None:
                unique_key_list = []
                if isinstance(unique_key, str) or isinstance(unique_key, unicode):
                    if unique_key in columns:
                        unique_key_list.append(unique_key)
                    else:
                        unique_key = None
                elif isinstance(unique_key, list) or isinstance(unique_key, tuple):
                    update_column = columns[:]
                    for ukey in unique_key:
                        if ukey in columns:
                            unique_key_list.append(ukey)
                            update_column.remove(ukey)
                    if len(unique_key_list) == 0:
                        unique_key = None
                    else:
                        values_occup = map(lambda x: '%s=%%s' % x, update_column)
                        update_string = 'UPDATE %s SET %s WHERE ' % (tb_name, ','.join(values_occup))

            for single_vals in values:
                insert_list = []
                update_list = []
                # upsert_list = []

                for i in range(0, len(columns)):
                    temp_value = single_vals[i]
                    if isinstance(temp_value, str):
                        temp_value.replace('%', '')
                        temp_value = MySQLdb.escape_string(temp_value)
                    elif isinstance(temp_value, unicode):
                        temp_value.replace('%', '')
                        temp_value = MySQLdb.escape_string(temp_value.encode('utf-8', 'ignore'))
                    insert_list.append(temp_value)
                    if unique_key is not None:
                        if columns[i] not in unique_key_list:
                            update_list.append(temp_value)
                    # upsert_list.append('`%s` = "%s"' % (columns[i], temp_value))

                try:
                    if unique_key is not None:
                        condition_list = []
                        for ukey in unique_key_list:
                            uindex = columns.index(ukey)
                            condition_list.append('`%s` = "%s"' % (ukey, insert_list[uindex]))

                        select_string = 'SELECT count(*) FROM %s WHERE %s;' % (tb_name, ' AND '.join(condition_list))
                        self.cursor.execute(select_string)
                        res_count = self.cursor.fetchone()

                        if int(res_count[0]) == 0:
                            self.cursor.execute(insert_string, insert_list)
                            self.conn.commit()
                        else:
                            update_query = '%s %s;' % (update_string, ' AND '.join(condition_list))
                            self.cursor.execute(update_query, update_list)
                            self.conn.commit()
                    else:
                        self.cursor.execute(insert_string, insert_list)
                        self.conn.commit()
                    # # update and insert
                    # upsert_string = '%s %s;' % (upsert_string, ','.join(upsert_list))
                    # self.cursor.execute(upsert_string, insert_list)
                except Exception as what:
                    print what
                    # self.conn.rollback()
        self.releaseConn()

    def mysqlDestroy(self):
        try:
            self.cursor.close()
            self.conn.close()
            print 'Mysql destroyed %s %s' % (self._host, self._db)
        except MySQLdb.Error,e:
            print 'Mysql Error %d: %s' % (e.args[0], e.args[1])

if __name__ == '__main__':
    columns = ['a', 'b', 'c']
    values = [['za',23, ''], ['as',32, '']]
    column_list = map(lambda x: MySQLdb.escape_string(x), columns)
    print ','.join(column_list)
